
from __future__ import annotations
import json

# ========== Standard / Django imports ==========
import io
import os
import tempfile
import calendar
from decimal import Decimal

import pdfkit

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import models
from django.db.models import Q, Sum, Count
from django.forms.widgets import NumberInput
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template, render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.formats import date_format
from django.utils.timezone import localtime, now
from django.utils.translation import activate
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.core.mail import EmailMessage
from .utils import build_email

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet

from weasyprint import HTML, CSS

# ========== datetime strategy ==========
import datetime as dt_module
from datetime import datetime as dt, timedelta, date

# ========== Optional Excel export dependency ==========
try:
    import openpyxl  # type: ignore
    # Styles will be imported within functions when needed
except Exception:  # ImportError or any environment issue
    openpyxl = None  # type: ignore

# ========== Local app imports ==========
from .models import (
    Facturation, Code, Paiement,
    PrevisionHospitalisation, Message, Observation, Patient, Courrier, CourrierPhoto,
)
from .forms import FacturationForm, PrevisionHospitalisationForm, ObservationForm, PatientForm, CourrierForm


# ========== Utilities ==========
def parse_flex_date(s: str | None) -> dt_module.date | None:
    """
    Essaie de parser une date aux formats 'YYYY-MM-DD' ou 'DD/MM/YYYY'.
    Retourne un objet date (naive) ou None si invalide.
    """
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ========== Facturation: Search / CRUD / Detail ==========

class FacturationSearchListView(LoginRequiredMixin, ListView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    template_name = 'comptabilite/facturation_search_list.html'
    context_object_name = 'facturations'

    def get_queryset(self):
        import re
        qs = super().get_queryset()

        q = (self.request.GET.get('q') or '').strip()
        if q:
            # Recherche texte habituelle
            text_q = (
                Q(dn__icontains=q) |
                Q(nom__icontains=q) |
                Q(prenom__icontains=q) |
                Q(numero_facture__icontains=q) |
                Q(code_acte__code_acte__icontains=q)
            )

            # Recherche date intelligente
            date_q = Q()
            # 1) Jour exact (JJ/MM/AAAA ou AAAA-MM-JJ)
            d = parse_flex_date(q)
            if d:
                date_q |= Q(date_acte=d)

            # 2) Mois/Année (MM/AAAA)
            m = re.fullmatch(r'(?P<mois>0[1-9]|1[0-2])/(?P<an>\d{4})', q)
            if m:
                an = int(m.group('an'))
                mois = int(m.group('mois'))
                from calendar import monthrange
                first_day = date(an, mois, 1)
                last_day = date(an, mois, monthrange(an, mois)[1])
                date_q |= Q(date_acte__range=(first_day, last_day))

            # 3) Mois/Année (AAAA-MM)
            m = re.fullmatch(r'(?P<an>\d{4})-(?P<mois>0[1-9]|1[0-2])', q)
            if m:
                an = int(m.group('an'))
                mois = int(m.group('mois'))
                from calendar import monthrange
                first_day = date(an, mois, 1)
                last_day = date(an, mois, monthrange(an, mois)[1])
                date_q |= Q(date_acte__range=(first_day, last_day))

            # 4) Année (AAAA)
            if re.fullmatch(r'\d{4}', q):
                date_q |= Q(date_acte__year=int(q))

            qs = qs.filter(text_q | date_q)

        # Filtres période existants
        today = timezone.localdate()
        if self.request.GET.get('today'):
            qs = qs.filter(date_acte=today)

        if self.request.GET.get('week'):
            start_of_week = today - timedelta(days=today.weekday())
            end_of_week   = start_of_week + timedelta(days=6)
            qs = qs.filter(date_acte__gte=start_of_week, date_acte__lte=end_of_week)

        if self.request.GET.get('month'):
            first_of_month = today.replace(day=1)
            from calendar import monthrange
            last_of_month  = date(today.year, today.month, monthrange(today.year, today.month)[1])
            qs = qs.filter(date_acte__gte=first_of_month, date_acte__lte=last_of_month)

        return qs.order_by('-date_acte')


class FacturationListView(LoginRequiredMixin, ListView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    template_name = 'comptabilite/facturation_list.html'
    context_object_name = 'facturations'

    def get_queryset(self):
        qs = super().get_queryset()
        # Si on a coché la case “today” dans le GET, on ne garde que les factures du jour
        if self.request.GET.get('today'):
            today = timezone.localdate()
            qs = qs.filter(date_acte=today)
        return qs


class FacturationCreateView(LoginRequiredMixin, CreateView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    form_class = FacturationForm
    template_name = 'comptabilite/facturation_form.html'
    success_url = reverse_lazy('facturation_list')

    def get_initial(self):
        initial = super().get_initial()
        # Pré-remplir depuis les paramètres GET (depuis la fiche patient)
        for key in ('dn', 'nom', 'prenom', 'date_naissance'):
            val = self.request.GET.get(key)
            if val:
                initial[key] = val
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Charger tous les codes et préparer un dictionnaire avec les données à transmettre
        codes = Code.objects.all()
        codes_data = {}
        for code in codes:
            codes_data[code.code_acte] = {
                'total_acte': str(code.total_acte),
                'tiers_payant': str(code.tiers_payant),
                'total_paye': str(code.total_paye),
            }
        context['codes_data'] = json.dumps(codes_data)
        return context

    def form_valid(self, form):
        inst = form.instance
        if (inst.lieu_acte or '').lower() == 'clinique':
            inst.numero_facture = ''  # force vide
        return super().form_valid(form)


class FacturationUpdateView(LoginRequiredMixin, UpdateView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    form_class = FacturationForm
    template_name = 'comptabilite/facturation_form.html'
    success_url = reverse_lazy('facturation_list')

    def form_valid(self, form):
        inst = form.instance
        if (inst.lieu_acte or '').lower() == 'clinique':
            inst.numero_facture = ''  # force vide
        return super().form_valid(form)


class FacturationDeleteView(LoginRequiredMixin, DeleteView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    template_name = 'comptabilite/facturation_confirm_delete.html'
    success_url = reverse_lazy('facturation_list')


class FacturationDetailView(LoginRequiredMixin, DetailView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    template_name = 'comptabilite/facturation_detail.html'
    context_object_name = 'facturation'


def check_dn(request):
    dn = request.GET.get('dn')
    if dn:
        try:
            fact = Facturation.objects.filter(dn=dn).latest('id')
            data = {
                'dn': fact.dn,
                'nom': fact.nom,
                'prenom': fact.prenom,
                'date_naissance': fact.date_naissance.strftime('%Y-%m-%d') if fact.date_naissance else "",
            }
            return JsonResponse({'exists': True, 'patient': data})
        except Facturation.DoesNotExist:
            pass
    return JsonResponse({'exists': False})


def check_acte(request):
    code_value = request.GET.get('code')
    if code_value:
        try:
            # Récupère l'objet Code par sa clé primaire
            code_obj = Code.objects.get(pk=code_value)
            data = {
                'total_acte': str(int(round(code_obj.total_acte))),
                'tiers_payant': str(int(round(code_obj.tiers_payant))),
                'total_paye': str(int(round(code_obj.total_paye))),
            }
            return JsonResponse({'exists': True, 'data': data})
        except Code.DoesNotExist:
            return JsonResponse({'exists': False})
    return JsonResponse({'exists': False})


def print_facture(request, pk):
    facture = get_object_or_404(Facturation, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="facture_{}.pdf"'.format(facture.numero_facture or facture.pk)
    c = canvas.Canvas(response, pagesize=A4)
    largeur, hauteur = A4

    # Récupération des données
    nom = facture.nom
    prenom = facture.prenom
    date_naissance = facture.date_naissance.strftime("%d/%m/%Y") if facture.date_naissance else ""
    dn = facture.dn
    date_facture = facture.date_facture.strftime("%d/%m/%Y") if facture.date_facture else ""
    total_acte = str(facture.total_acte)
    total_paye = str(facture.total_paye)
    tiers_payant = str(facture.tiers_payant)

    code_obj = facture.code_acte
    if code_obj:
        code_acte_normal = code_obj.code_acte_normal or ""
        variable_1 = code_obj.variable_1 or ""
        modificateur = code_obj.modificateur or ""
        variable_2 = code_obj.variable_2 or ""
        ps = "X" if code_obj.parcours_soin else ""
    else:
        code_acte_normal = ""
        variable_1 = ""
        modificateur = ""
        variable_2 = ""
        ps = ""

    if code_obj and code_obj.medecin:
        nom_medecin = code_obj.medecin.nom_medecin
        nom_clinique = code_obj.medecin.nom_clinique
        code_m = code_obj.medecin.code_m
    else:
        nom_medecin = ""
        nom_clinique = ""
        code_m = ""

    regime_lm = ""
    if hasattr(facture, 'regime_lm'):
        regime_lm = "X" if facture.regime_lm else ""

    # Si hors parcours de soins ET pas encore de numéro, on le génère et on enregistre
    if (
        not facture.code_acte.parcours_soin
        and not facture.numero_facture
        and (getattr(facture, 'lieu_acte', '') or '').lower() != 'clinique'
    ):
        now_local = timezone.localtime()
        facture.numero_facture = now_local.strftime("JA/%Y/%m/%d/%H:%M")
        facture.save()

    c.drawString(2.0 * cm, hauteur - 3.7 * cm, f"{nom}")
    c.drawString(12.5 * cm, hauteur - 3.7 * cm, f"{prenom}")
    c.drawString(16.0 * cm, hauteur - 4.7 * cm, f"{date_naissance}")
    c.drawString(2.0 * cm, hauteur - 4.7 * cm, f"{dn}")
    c.drawString(2.0 * cm, hauteur - 13.0 * cm, f"{nom_medecin}")
    c.drawString(2.0 * cm, hauteur - 13.5 * cm, f"{nom_clinique}")
    c.drawString(10.5 * cm, hauteur - 12.6 * cm, f"{code_m}")
    c.drawString(2.7 * cm, hauteur - 14.7 * cm, f"{ps}")
    if facture.regime_lm:
        c.drawString(0.3 * cm, hauteur - 16.5 * cm, f"{regime_lm}")
    c.drawString(0.5 * cm, hauteur - 20.3 * cm, f"{date_facture}")
    c.drawString(4.0 * cm, hauteur - 20.3 * cm, f"{code_acte_normal}")
    c.drawString(7.0 * cm, hauteur - 20.3 * cm, f"{variable_1}")
    c.drawString(7.5 * cm, hauteur - 20.3 * cm, f"{modificateur}")
    c.drawString(11.5 * cm, hauteur - 20.3 * cm, f"{variable_2}")
    c.drawString(12.5 * cm, hauteur - 20.3 * cm, f"{total_acte}")
    c.drawString(10.0 * cm, hauteur - 24.3 * cm, f"{total_acte}")
    if facture.regime_lm or (facture.tiers_payant and facture.tiers_payant != Decimal('0')):
        c.drawString(7.5 * cm, hauteur - 27.6 * cm, f"{total_paye}")
        c.drawString(11.5 * cm, hauteur - 27.6 * cm, f"{tiers_payant}")

    c.save()
    return response


# ========== Bordereau ==========

@login_required
def create_bordereau(request):
    """
    Aperçu des factures à bordereauter, triées par numero_facture,
    sans rien modifier en base.
    """
    factures = (
        Facturation.objects
        .filter(tiers_payant__gt=0)
        .filter(Q(numero_bordereau__isnull=True) | Q(numero_bordereau=""))
        .order_by('numero_facture')
    )

    if not factures.exists():
        return render(request, 'comptabilite/bordereau.html', {
            'error': "Aucune facture à traiter pour le bordereau."
        })

    today_local = date.today()
    week = today_local.isocalendar()[1]
    day_of_year = today_local.timetuple().tm_yday
    num_bordereau = f"M{today_local.year}-{today_local.month:02d}-{week:02d}-{day_of_year:03d}"

    context = {
        'factures': factures,
        'num_bordereau': num_bordereau,
        'date_bordereau': today_local.strftime("%d/%m/%Y"),
        'count': factures.count(),
        'total_tiers_payant': factures.aggregate(total=Sum('tiers_payant'))['total'] or 0,
    }
    return render(request, 'comptabilite/bordereau.html', context)


@login_required
def print_bordereau(request, num_bordereau):
    """
    Impression du bordereau :
    - On sélectionne les factures non encore bordereautées.
    - On les matérialise dans une liste pour l’affichage.
    - On les met à jour en base.
    - On génère le PDF à partir de la liste déjà chargée.
    """
    # 1) QuerySet initial
    qs = (Facturation.objects
          .filter(tiers_payant__gt=0)
          .filter(Q(numero_bordereau__isnull=True) | Q(numero_bordereau=""))
          .order_by('numero_facture'))

    if not qs.exists():
        return HttpResponse("Aucune facture à imprimer pour ce bordereau.", status=404)

    # 2) On matérialise avant update
    factures_list = list(qs)

    # 3) Totaux
    count = len(factures_list)
    total_tiers = sum(f.tiers_payant for f in factures_list)

    # 4) Mise à jour en base
    today_local = date.today()
    qs.update(numero_bordereau=num_bordereau, date_bordereau=today_local)

    # 5) Contexte, on passe la liste « figée »
    context = {
        'num_bordereau':      num_bordereau,
        'date_bordereau':     today_local.strftime("%d/%m/%Y"),
        'count':              count,
        'total_tiers_payant': total_tiers,
        'factures':           factures_list,
    }

    # 6) Génération du PDF
    html_string = render_to_string('comptabilite/bordereau_pdf.html', context)
    pdf_file    = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{num_bordereau}.pdf"'
    return response


# ========== Activity list / filters ==========

class ActivityListView(LoginRequiredMixin, ListView):
    login_url = 'login'
    redirect_field_name = 'next'   # paramètre renvoyé après login
    model = Facturation
    template_name = 'comptabilite/activity_list.html'
    context_object_name = 'factures'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Récupérer les paramètres de filtre
        date_str = self.request.GET.get('date')          # 'YYYY-MM-DD' ou 'DD/MM/YYYY'
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        year_str = self.request.GET.get('year')            # ex: '2025'

        if date_str:
            filter_date = parse_flex_date(date_str)
            if filter_date:
                queryset = queryset.filter(date_facture=filter_date)

        elif start_date_str and end_date_str:
            start_date = parse_flex_date(start_date_str)
            end_date   = parse_flex_date(end_date_str)
            if start_date and end_date:
                queryset = queryset.filter(date_facture__gte=start_date,
                                           date_facture__lte=end_date)

        elif year_str:
            # Filtrer par année
            try:
                year = int(year_str)
                queryset = queryset.filter(date_facture__year=year)
            except ValueError:
                pass

        return queryset.order_by('date_facture')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['date'] = self.request.GET.get('date', '')
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['year'] = self.request.GET.get('year', '')
        aggregates = self.get_queryset().aggregate(
            sum_total_acte=Sum('total_acte'),
            sum_tiers_payant=Sum('tiers_payant'),
            sum_total_paye=Sum('total_paye')
        )
        context['sum_total_acte'] = aggregates['sum_total_acte'] or 0
        context['sum_tiers_payant'] = aggregates['sum_tiers_payant'] or 0
        context['sum_total_paye'] = aggregates['sum_total_paye'] or 0
        return context


# ========== Chèques (listing / choix date / remise / impression) ==========

@login_required
def cheque_listing(request):
    # POST : on coche tous les chèques non listés dont la date ≤ aujourd'hui
    if request.method == 'POST':
        Paiement.objects.filter(
            modalite_paiement="Chèque",
            date__lte=date.today(),
            liste=False
        ).update(liste=True)
        return redirect('cheque_listing')

    # GET : on n'affiche que ceux qui ne sont pas encore listés
    cheques = Paiement.objects.filter(
        modalite_paiement="Chèque",
        date__lte=date.today(),
        liste=False
    )
    count = cheques.count()
    total_montant = cheques.aggregate(total=Sum('montant'))['total'] or 0

    return render(request, 'comptabilite/cheque_listing.html', {
        'cheques': cheques,
        'count': count,
        'total_montant': total_montant,
        'filter_date': date.today().strftime("%d/%m/%Y"),
    })


@login_required
def choix_date_cheques(request):
    """
    Permet de choisir la date pour la remise des chèques.
    Si l'utilisateur sélectionne "Aujourd'hui", on redirige avec la date du jour.
    Sinon, on utilise la date saisie.
    """
    if request.method == 'POST':
        option = request.POST.get('date_option')
        if option == 'today':
            chosen_date = date.today().strftime('%Y-%m-%d')
            return redirect(f"/facturation/cheques/?date={chosen_date}")
        elif option == 'other':
            other_date = request.POST.get('other_date')
            if other_date:
                parsed = parse_flex_date(other_date)
                if parsed:
                    return redirect(f"/facturation/cheques/?date={other_date}")
                else:
                    context = {'error': "Format de date invalide. Utilisez AAAA-MM-JJ ou JJ/MM/AAAA."}
                    return render(request, 'comptabilite/choix_date_cheques.html', context)
    return render(request, 'comptabilite/choix_date_cheques.html', {})


class IntegerNumberInput(NumberInput):
    def format_value(self, value):
        if value is None or value == '':
            return ''
        try:
            return str(int(round(float(value))))
        except (ValueError, TypeError):
            return super().format_value(value)


@login_required
def remise_cheque(request):
    """
    Affiche la liste des paiements (chèques) non listés et à date <= aujourd'hui,
    puis, quand on valide (POST), marque 'listé' = True et renvoie le PDF.
    """
    today_local = timezone.localdate()

    if request.method == 'POST':
        cheques = Paiement.objects.filter(
            modalite_paiement='Chèque',
            date__lte=today_local,
            liste=False
        )
        cheques.update(liste=True)

        count = cheques.count()
        total = cheques.aggregate(total=models.Sum('montant'))['total'] or 0

        context = {
            'cheques': cheques,
            'count': count,
            'date_cheque': today_local.strftime('%d/%m/%Y'),
            'total_montant': total,
        }

        html = render_to_string('comptabilite/remise_cheque_pdf.html', context)
        pdf = pdfkit.from_string(html, False, options=getattr(settings, 'PDFKIT_OPTIONS', {}))

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="remise_cheque_{today_local}.pdf"'
        return response

    cheques = Paiement.objects.filter(
        modalite_paiement='Chèque',
        date__lte=today_local,
        liste=False
    )
    count = cheques.count()
    total = cheques.aggregate(total=models.Sum('montant'))['total'] or 0

    return render(request, 'comptabilite/remise_cheque.html', {
        'cheques': cheques,
        'count': count,
        'date_cheque': today_local.strftime('%d/%m/%Y'),
        'total_montant': total,
    })


@login_required
def print_cheque_listing(request):
    # 1. Récupération de la date
    d_str = request.GET.get('date')
    if d_str:
        filter_date = parse_flex_date(d_str) or date.today()
    else:
        filter_date = date.today()

    # 2. Requête initiale (liste=False)
    qs = Paiement.objects.filter(
        modalite_paiement="Chèque",
        date__lte=filter_date,
        liste=False
    ).order_by('date')

    # 3. On cache en mémoire AVANT le .update()
    cheques = list(qs)
    if not cheques:
        return redirect('cheque_listing')

    count = len(cheques)
    total_montant = sum(ch.montant for ch in cheques)

    # 4. On marque TOUTES ces lignes comme « listées »
    qs.update(liste=True)

    # 5. Construction du PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - 2*cm
    c.setFont('Helvetica-Bold', 16)
    c.drawString(2*cm, y, "Dr. Jean-Ariel BRONSTEIN | Gastroentérologue")
    y -= 1*cm

    c.setFont('Helvetica-Bold', 16)
    c.drawString(2*cm, y, "Rue Lagarde | Papeete")
    y -= 1*cm

    c.setFont('Helvetica-Bold', 16)
    c.drawString(2*cm, y, f"Remise de {count} Chèque(s)")
    y -= 1*cm

    c.setFont('Helvetica', 12)
    c.drawString(2*cm, y, f"Date de remise : {filter_date.strftime('%d/%m/%Y')}")
    y -= 1*cm

    # En‑tête du tableau
    c.setFont('Helvetica-Bold', 12)
    xs = [2*cm, 6*cm, 10*cm, 14*cm]
    for x, title in zip(xs, ["Date", "Banque", "Porteur", "Montant"]):
        c.drawString(x, y, title)
    y -= 0.7*cm

    # Lignes détaillées
    c.setFont('Helvetica', 11)
    for ch in cheques:
        if y < 2*cm:
            c.showPage()
            y = height - 2*cm
            # on ré‑affiche l’en‑tête
            c.setFont('Helvetica-Bold', 12)
            for x, title in zip(xs, ["Date", "Banque", "Porteur", "Montant"]):
                c.drawString(x, y, title)
            y -= 0.7*cm
            c.setFont('Helvetica', 11)

        c.drawString(xs[0], y, ch.date.strftime('%d/%m/%Y'))
        c.drawString(xs[1], y, ch.banque or "")
        c.drawString(xs[2], y, ch.porteur or "")
        c.drawRightString(xs[3] + 3*cm, y, f"{int(ch.montant)}")
        y -= 0.7*cm

    # Total en bas
    y -= 1*cm
    c.setFont('Helvetica-Bold', 12)
    c.drawString(2*cm, y, f"Nombre de chèque(s) : {count}   Total des montants : {int(total_montant)}")

    c.save()
    buffer.seek(0)

    # 6. Retour HTTP
    filename = f"remise_cheque_{filter_date.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ========== Numéro facture util ==========

def generate_numero(request, pk):
    """
    Génère et sauvegarde un numero_facture pour la Facturation pk
    si elle n'en a pas encore et que parcours_soin == False (et pas Clinique).
    """
    facture = get_object_or_404(Facturation.objects.select_related('code_acte'), pk=pk)
    code = facture.code_acte
    if (
        not facture.numero_facture
        and not (code and code.parcours_soin)
        and (getattr(facture, 'lieu_acte', '') or '').lower() != 'clinique'
    ):
        now_local = timezone.localtime()
        facture.numero_facture = (
            f"JA/{now_local.year}/{now_local.month:02d}/{now_local.day:02d}/"
            f"{now_local.hour:02d}:{now_local.minute:02d}"
        )
        facture.save(update_fields=['numero_facture'])
    return JsonResponse({'numero_facture': facture.numero_facture or ""})


# ========== Comptabilité: summaries / pivots ==========

class ComptabiliteSummaryView(LoginRequiredMixin, ListView):
    login_url = 'login'
    redirect_field_name = 'next'
    model = Facturation
    template_name = 'comptabilite/comptabilite_summary.html'
    context_object_name = 'summary_rows'

    def get_queryset(self):
        qs = super().get_queryset()
        period = self.request.GET.get('period', '')
        today_local = timezone.localdate()

        if period == 'today':
            qs = qs.filter(date_facture=today_local)
        elif period == 'week':
            start = today_local - timedelta(days=today_local.weekday())
            end = start + timedelta(days=6)
            qs = qs.filter(date_facture__range=(start, end))
        elif period == 'month':
            qs = qs.filter(date_facture__year=today_local.year, date_facture__month=today_local.month)
        elif period == 'year':
            qs = qs.filter(date_facture__year=today_local.year)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()

        ctx['period_choices'] = [
            ('', 'Toutes'),
            ('today', "Aujourd'hui"),
            ('week', 'Cette semaine'),
            ('month', 'Ce mois'),
            ('year', "Cette année"),
        ]
        ctx['period'] = self.request.GET.get('period', '')

        ctx['group_regime']    = bool(self.request.GET.get('group_regime'))
        ctx['group_modalite']  = bool(self.request.GET.get('group_modalite'))
        ctx['group_code_reel'] = bool(self.request.GET.get('group_code_reel'))
        ctx['group_lieu']      = bool(self.request.GET.get('group_lieu'))

        # Pivot par régime
        if ctx['group_regime']:
            rows = (qs.values('regime')
                     .annotate(
                         count=Count('id'),
                         total_acte=Sum('total_acte'),
                         total_paye=Sum('total_paye')
                     ))
        else:
            rows = []
        ctx['pivot_regime'] = rows
        ctx['totals_regime'] = {
            'count': sum(r['count'] for r in rows),
            'total_acte': sum(r['total_acte'] or 0 for r in rows),
            'total_paye': sum(r['total_paye'] or 0 for r in rows),
        }

        # Pivot par modalité
        if ctx['group_modalite']:
            tmp = (qs.select_related('paiement')
                     .values('paiement__modalite_paiement')
                     .annotate(
                         count=Count('id'),
                         total_acte=Sum('total_acte'),
                         total_paye=Sum('total_paye')
                     )
                     .order_by('paiement__modalite_paiement'))
            rows_mod = [
                {'label': r['paiement__modalite_paiement'], **r}
                for r in tmp
            ]
        else:
            rows_mod = []
        ctx['pivot_modalite'] = rows_mod
        ctx['totals_modalite'] = {
            'count': sum(r['count'] for r in rows_mod),
            'total_acte': sum(r['total_acte'] or 0 for r in rows_mod),
            'total_paye': sum(r['total_paye'] or 0 for r in rows_mod),
        }

        # Pivot par code réel
        if ctx['group_code_reel']:
            tmp = (qs.select_related('code_acte')
                     .values('code_acte__code_reel')
                     .annotate(
                         count=Count('id'),
                         total_acte=Sum('total_acte'),
                         total_paye=Sum('total_paye')
                     )
                     .order_by('code_acte__code_reel'))
            rows_code = [
                {'code_reel': r['code_acte__code_reel'], **r}
                for r in tmp
            ]
        else:
            rows_code = []
        ctx['pivot_code_reel'] = rows_code
        ctx['totals_code_reel'] = {
            'count': sum(r['count'] for r in rows_code),
            'total_acte': sum(r['total_acte'] or 0 for r in rows_code),
            'total_paye': sum(r['total_paye'] or 0 for r in rows_code),
        }

        # Pivot par lieu d'acte
        if ctx['group_lieu']:
            tmp = (qs.values('lieu_acte')
                     .annotate(
                         count=Count('id'),
                         total_acte=Sum('total_acte'),
                         total_paye=Sum('total_paye')
                     )
                     .order_by('lieu_acte'))
            rows_lieu = [
                {'lieu_acte': r['lieu_acte'], **r}
                for r in tmp
            ]
        else:
            rows_lieu = []
        ctx['pivot_lieu'] = rows_lieu
        ctx['totals_lieu'] = {
            'count': sum(r['count'] for r in rows_lieu),
            'total_acte': sum(r['total_acte'] or 0 for r in rows_lieu),
            'total_paye': sum(r['total_paye'] or 0 for r in rows_lieu),
        }

        return ctx


# ========== Prévisions d’hospitalisation (CRUD + PDF + Email) ==========

def prevision_list(request):
    previsions = PrevisionHospitalisation.objects.all().order_by('-date_entree')
    today_local = localtime().date()

    return render(request, 'comptabilite/prevision_list.html', {
        'previsions': previsions,
        'today': today_local,
    })


def prevision_create(request):
    form = PrevisionHospitalisationForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('prevision_list')
    return render(request, 'comptabilite/prevision_form.html', {'form': form})


def prevision_detail(request, pk):
    prevision = get_object_or_404(PrevisionHospitalisation, pk=pk)
    return render(request, 'comptabilite/prevision_detail.html', {'prevision': prevision})


def prevision_pdf(request, pk):
    prevision = get_object_or_404(PrevisionHospitalisation, pk=pk)
    template = get_template('comptabilite/prevision_detail.html')
    logo_abspath = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')

    html_content = template.render({
        'prevision': prevision,
        'logo_path': logo_abspath,
    })

    css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')

    with tempfile.NamedTemporaryFile(delete=True, suffix=".pdf") as output:
        HTML(string=html_content, base_url=request.build_absolute_uri()).write_pdf(
            output.name,
            stylesheets=[CSS(filename=css_path)]
        )
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="prevision_{prevision.pk}.pdf"'
        return response


def prevision_send_email(request, pk):
    prevision = get_object_or_404(PrevisionHospitalisation, pk=pk)

    # chemins vers logo et CSS
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
    css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')

    # rendu HTML
    template = get_template('comptabilite/prevision_detail.html')
    html_content = template.render({'prevision': prevision, 'logo_path': logo_path})

    # ✅ destinataires selon le lieu d’hospitalisation
    if prevision.lieu_hospitalisation == "Médecine":
        destinataires = [
            "ejosse@polyclinique-paofai.pf",
            "secretariat@bronstein.fr",
            "medecine@polyclinique-paofai.pf",
            "docteur@bronstein.fr"
        ]
    else:
        destinataires = [
            "anesthesiepaofai@gmail.com",
            "ebarce@polyclinique-paofai.pf",
            "bronstein.tahiti@proton.me",
            "secretariat@bronstein.fr",
            "docteur@bronstein.fr"
        ]

    # génération du PDF
    with tempfile.NamedTemporaryFile(delete=True, suffix=".pdf") as pdf_file:
        HTML(string=html_content, base_url=request.build_absolute_uri()).write_pdf(
            pdf_file.name, stylesheets=[CSS(filename=css_path)]
        )
        pdf_file.seek(0)

        # composition du mail (centralisé)
        email = build_email(
            subject=f"Prévision d’hospitalisation – {prevision.nom} {prevision.prenom} {prevision.date_naissance}",
            body="Bonjour,\n\nVeuillez trouver ci-joint la fiche de prévision d'hospitalisation.\n\nBien cordialement,\nDr. Bronstein",
            to=destinataires,
        )
        email.attach(f"prevision_{prevision.pk}.pdf", pdf_file.read(), 'application/pdf')
        email.send()

    return redirect('prevision_detail', pk=prevision.pk)


def prevision_update(request, pk):
    prevision = get_object_or_404(PrevisionHospitalisation, pk=pk)
    form = PrevisionHospitalisationForm(request.POST or None, instance=prevision)
    if form.is_valid():
        form.save()
        return redirect('prevision_list')
    return render(request, 'comptabilite/prevision_form.html', {'form': form})


def prevision_delete(request, pk):
    prevision = get_object_or_404(PrevisionHospitalisation, pk=pk)
    if request.method == 'POST':
        prevision.delete()
        return redirect('prevision_list')
    return render(request, 'comptabilite/prevision_confirm_delete.html', {'prevision': prevision})


# ========== Mini chat ==========

@login_required
def chat_view(request, receiver_id):
    receiver = get_object_or_404(User, pk=receiver_id)
    return render(request, 'comptabilite/chat.html', {'receiver': receiver})


@login_required
def get_messages(request):
    receiver_id = request.GET.get('receiver_id')
    messages = Message.objects.filter(
        sender=request.user, receiver_id=receiver_id
    ) | Message.objects.filter(
        sender_id=receiver_id, receiver=request.user
    )
    # marquer les messages reçus comme lus
    Message.objects.filter(sender_id=receiver_id, receiver=request.user, lu=False).update(lu=True)

    data = [{
        'sender_id': m.sender.id,
        'sender_name': m.sender.get_full_name() or m.sender.username,
        'content': m.content,
        'timestamp': m.timestamp.isoformat(),
        'lu': m.lu
    } for m in messages]

    return JsonResponse({'messages': data})


@login_required
def send_message(request):
    if request.method == 'POST':
        content = request.POST.get('content')
        receiver_id = request.POST.get('receiver_id')
        Message.objects.create(
            sender=request.user,
            receiver_id=receiver_id,
            content=content
        )
        return JsonResponse({'status': 'ok'})


# ========== Patients hospitalisés (liste / PDF / Excel) ==========

@login_required
def patients_hospitalises(request):
    context = get_patients_hospitalises()
    return render(request, "comptabilite/patients_hospitalises.html", context)


@login_required
def patients_hospitalises_pdf(request):
    context = get_patients_hospitalises()
    context['user'] = request.user

    html_string = render_to_string("comptabilite/patients_hospitalises_pdf.html", context)
    css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')

    base_url = request.build_absolute_uri('/') if settings.DEBUG else f'file://{settings.BASE_DIR}/'
    pdf = HTML(string=html_string, base_url=base_url).write_pdf(
        stylesheets=[CSS(filename=css_path)]
    )

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="patients_hospitalises.pdf"'
    return response


def get_patients_hospitalises():
    today_local = now().date()
    qs = PrevisionHospitalisation.objects.filter(
        date_entree__lte=today_local
    ).filter(
        Q(date_sortie__gte=today_local) | Q(date_sortie__isnull=True)
    )
    return {
        "today_formatted": today_local,  # laisser le formatage à Django
        "bloc": qs.filter(lieu_hospitalisation='Bloc'),
        "medecine": qs.filter(lieu_hospitalisation='Médecine'),
        "soins_continus": qs.filter(lieu_hospitalisation='Soins continus'),
    }


@login_required
def patients_hospitalises_excel(request):
    today_local = timezone.localdate()
    hospitalises = PrevisionHospitalisation.objects.filter(
        date_entree__lte=today_local
    ).filter(Q(date_sortie__gte=today_local) | Q(date_sortie__isnull=True))
    if openpyxl is None:
        return HttpResponse("openpyxl non installé – export Excel indisponible.", status=500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients hospitalisés"

    headers = ['DN', 'Nom', 'Prénom', 'Date Naissance', 'Entrée', 'Sortie Théorique', 'Sortie', 'Lieu']
    ws.append(headers)

    for p in hospitalises:
        sortie_theorique = p.date_sortie or (p.date_entree + dt_module.timedelta(days=1))
        ws.append([
            p.dn, p.nom, p.prenom,
            p.date_naissance.strftime('%d/%m/%Y') if p.date_naissance else '',
            p.date_entree.strftime('%d/%m/%Y') if p.date_entree else '',
            sortie_theorique.strftime('%d/%m/%Y') if sortie_theorique else '',
            p.date_sortie.strftime('%d/%m/%Y') if p.date_sortie else '',
            p.lieu_hospitalisation
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=patients_hospitalises_{today_local}.xlsx'
    wb.save(response)
    return response


# ========== Impression fiche facturation (PDF) ==========

@login_required
def imprimer_fiche_facturation(request, pk):
    activate('fr')  # Pour forcer les dates au format français

    facture = get_object_or_404(Facturation, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="facture_{facture.numero_facture or facture.pk}.pdf"'
    c = canvas.Canvas(response, pagesize=A4)
    largeur, hauteur = A4

    today_str = date_format(timezone.localdate(), format='d F Y', use_l10n=True)

    # En-tête
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, hauteur - 2 * cm, "Dr. Jean-Ariel BRONSTEIN")
    c.setFont("Helvetica", 12)
    c.drawString(2 * cm, hauteur - 2.7 * cm, "Gastro-entérologue")
    c.drawString(2 * cm, hauteur - 3.4 * cm, "Adresse : Clinique Paofai | Tel : 40.81.48.48")
    c.drawString(2 * cm, hauteur - 4.1 * cm, f"Date : {today_str}")

    # Corps de la facture
    c.setFont("Helvetica", 11)
    y = hauteur - 5.5 * cm
    lignes = [
        f"Facture n° : {facture.numero_facture or 'Non attribué'}",
        f"Patient : {facture.nom} {facture.prenom}",
        f"Date de naissance : {date_format(facture.date_naissance, 'd F Y') if facture.date_naissance else ''}",
        f"DN : {facture.dn}",
        f"Date de l'acte : {date_format(facture.date_acte, 'd F Y') if facture.date_acte else ''}",
        f"Date de la facture : {date_format(facture.date_facture, 'd F Y') if facture.date_facture else ''}",
        f"Code Acte : {facture.code_acte.code_acte if facture.code_acte else ''}",
        f"Montant total : {facture.total_acte} XPF",
        "",
        f"Payé le : {today_str}",
        "",
        "Signature : Dr. Jean-Ariel Bronstein"
    ]
    for line in lignes:
        c.drawString(2 * cm, y, line)
        y -= 1 * cm

    c.save()
    return response


# ========== Observations médicales (liste patients, CRUD, PDF) ==========

@login_required
def observations_patient_list(request):
    """Liste des patients ayant au moins une observation, avec recherche.

    Recherche sur: DN, nom, prénom, date de naissance (AAAA-MM-JJ ou JJ/MM/AAAA).
    """
    q = (request.GET.get('q') or '').strip()

    # Base: uniquement DNs présents dans Observation
    obs_base = Observation.objects.all()

    # Filtres de recherche
    if q:
        # Tente une date
        d = parse_flex_date(q)
        q_filter = (
            Q(dn__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q)
        )
        if d:
            q_filter |= Q(date_naissance=d)
        obs_base = obs_base.filter(q_filter)

    # Charger un sous-ensemble minimal pour agréger côté Python, en normalisant DN (trim)
    obs_rows = list(obs_base.values('dn', 'nom', 'prenom', 'date_naissance', 'date_observation'))

    # Dédoublonner par DN « normalisé » (trim) et calculer compte + dates
    dn_order = []               # ordre d'apparition
    obs_counts_map = {}         # dn_norm -> count
    obs_dates_map = {}          # dn_norm -> [dates]
    for r in obs_rows:
        dn_norm = (r['dn'] or '').strip()
        if not dn_norm:
            continue
        if dn_norm not in obs_counts_map:
            dn_order.append(dn_norm)
            obs_counts_map[dn_norm] = 0
            obs_dates_map[dn_norm] = []
        obs_counts_map[dn_norm] += 1
        d = r.get('date_observation')
        if d and d not in obs_dates_map[dn_norm]:
            obs_dates_map[dn_norm].append(d)

    # Trier les dates décroissantes pour l'affichage
    for dn_norm in obs_dates_map:
        obs_dates_map[dn_norm].sort(reverse=True)

    # Construire les lignes patient: on privilégie la dernière Facturation si dispo, sinon une Observation
    patients = []
    for dn in dn_order:
        fact = (
            Facturation.objects
            .filter(dn=dn)
            .order_by('-date_acte', '-id')
            .values('dn', 'nom', 'prenom', 'date_naissance')
            .first()
        )
        if fact:
            row = dict(fact)
        else:
            o = (
                Observation.objects
                .filter(dn=dn)
                .order_by('-date_observation', '-created_at')
                .values('dn', 'nom', 'prenom', 'date_naissance')
                .first()
            )
            row = dict(o) if o else {'dn': dn, 'nom': '', 'prenom': '', 'date_naissance': None}
        row['obs_count'] = obs_counts_map.get(dn, 0)
        row['obs_dates'] = obs_dates_map.get(dn, [])
        patients.append(row)

    # Tri par nom/prénom
    patients.sort(key=lambda p: (p.get('nom') or '', p.get('prenom') or ''))

    return render(request, 'comptabilite/observations_patient_list.html', {
        'patients': patients,
        'q': q,
    })


@login_required
def observations_by_dn(request, dn: str):
    """Liste des observations pour un DN, ordre chronologique."""
    observations = Observation.objects.filter(dn=dn).order_by('date_observation', 'created_at')

    # infos patient depuis Facturation si possible, sinon via une observation
    fact = (Facturation.objects.filter(dn=dn).order_by('-date_acte').first())
    if fact:
        patient = {
            'dn': fact.dn,
            'nom': fact.nom,
            'prenom': fact.prenom,
            'date_naissance': fact.date_naissance,
        }
    else:
        any_obs = observations.first()
        patient = {
            'dn': dn,
            'nom': any_obs.nom if any_obs else '',
            'prenom': any_obs.prenom if any_obs else '',
            'date_naissance': any_obs.date_naissance if any_obs else None,
        }

    return render(request, 'comptabilite/observations_list.html', {
        'patient': patient,
        'observations': observations,
    })


@login_required
def observation_create(request, dn: str | None = None):
    """Créer une observation. Pré-remplit depuis la dernière facturation si DN fourni."""
    initial = {}
    if dn:
        fact = Facturation.objects.filter(dn=dn).order_by('-date_acte').first()
        if fact:
            initial.update({
                'dn': fact.dn,
                'nom': fact.nom,
                'prenom': fact.prenom,
                'date_naissance': fact.date_naissance,
            })
        else:
            initial['dn'] = dn

    form = ObservationForm(request.POST or None, initial=initial)
    if form.is_valid():
        obs = form.save()
        return redirect('observations_by_dn', dn=obs.dn)

    return render(request, 'comptabilite/observation_form.html', {'form': form, 'mode': 'create'})


@login_required
def observation_update(request, pk: int):
    obs = get_object_or_404(Observation, pk=pk)
    form = ObservationForm(request.POST or None, instance=obs)
    if form.is_valid():
        obs = form.save()
        return redirect('observations_by_dn', dn=obs.dn)
    return render(request, 'comptabilite/observation_form.html', {'form': form, 'mode': 'update', 'obs': obs})


@login_required
def observation_delete(request, pk: int):
    obs = get_object_or_404(Observation, pk=pk)
    if request.method == 'POST':
        dn = obs.dn
        obs.delete()
        return redirect('observations_by_dn', dn=dn)
    return render(request, 'comptabilite/observation_confirm_delete.html', {'obs': obs})


@login_required
def observations_pdf(request, dn: str):
    """Génère une synthèse PDF des observations pour un patient (DN)."""
    observations = Observation.objects.filter(dn=dn).order_by('date_observation', 'created_at')
    fact = Facturation.objects.filter(dn=dn).order_by('-date_acte').first()
    patient = None
    if fact:
        patient = fact
    else:
        any_obs = observations.first()
        if any_obs:
            class P:
                pass
            patient = P()
            patient.dn = any_obs.dn
            patient.nom = any_obs.nom
            patient.prenom = any_obs.prenom
            patient.date_naissance = any_obs.date_naissance

    html_string = render_to_string('comptabilite/observations_pdf.html', {
        'patient': patient,
        'observations': observations,
        'user': request.user,
    })
    css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
        stylesheets=[CSS(filename=css_path)]
    )

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="observations_{dn}.pdf"'
    return response


@login_required
def observation_pdf(request, pk: int):
    """Génère un PDF pour une seule observation (par pk)."""
    obs = get_object_or_404(Observation, pk=pk)
    # Construire identité patient
    fact = Facturation.objects.filter(dn=obs.dn).order_by('-date_acte').first()
    class P:
        pass
    patient = P()
    if fact:
        patient.dn = fact.dn
        patient.nom = fact.nom
        patient.prenom = fact.prenom
        patient.date_naissance = fact.date_naissance
    else:
        patient.dn = obs.dn
        patient.nom = obs.nom
        patient.prenom = obs.prenom
        patient.date_naissance = obs.date_naissance

    html_string = render_to_string('comptabilite/observations_pdf.html', {
        'patient': patient,
        'observations': [obs],
        'user': request.user,
    })
    css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
        stylesheets=[CSS(filename=css_path)]
    )
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="observation_{pk}.pdf"'
    return response

@login_required
def observations_send_email(request, dn: str):
    """Affiche un petit formulaire pour envoyer le PDF des observations par e‑mail, puis envoie en pièce jointe."""
    # Construire l'identité patient (comme dans observations_pdf)
    observations = Observation.objects.filter(dn=dn).order_by('date_observation', 'created_at')
    fact = Facturation.objects.filter(dn=dn).order_by('-date_acte').first()
    if fact:
        patient = fact
    else:
        any_obs = observations.first()
        if any_obs:
            class P: pass
            patient = P()
            patient.dn = any_obs.dn
            patient.nom = any_obs.nom
            patient.prenom = any_obs.prenom
            patient.date_naissance = any_obs.date_naissance
        else:
            # Rien à envoyer
            return redirect('patients_detail', dn=dn)

    if request.method == 'POST':
        to = (request.POST.get('to') or '').strip()
        cc = (request.POST.get('cc') or '').strip()
        subject = (request.POST.get('subject') or '').strip() or f"Observations – {patient.nom} {patient.prenom} (DN {dn})"
        body = (request.POST.get('body') or '').strip() or "Bonjour,\n\nVeuillez trouver ci-joint le PDF des observations.\n\nBien cordialement,\nDr. Bronstein"
        to_list = [addr.strip() for addr in to.replace(';', ',').split(',') if addr.strip()]
        cc_list = [addr.strip() for addr in cc.replace(';', ',').split(',') if addr.strip()]

        if not to_list:
            messages.error(request, "Veuillez renseigner au moins un destinataire.")
            default_subject = f"Observations – {patient.nom} {patient.prenom} (DN {dn})"
            default_body = "Bonjour,\n\nVeuillez trouver ci-joint le PDF des observations.\n\nBien cordialement,\nDr. Bronstein"
            # Par défaut, on met le secrétariat en Cc et non en To, en évitant les doublons
            _cands = [request.user.email if getattr(request.user, 'email', '') else None, 'docteur@bronstein.fr']
            _cands = [x for x in _cands if x]
            default_to = ", ".join(list(dict.fromkeys([x.lower() for x in _cands])))
            return render(request, 'comptabilite/observations_email.html', {
                'patient': patient,
                'dn': dn,
                'default_subject': default_subject,
                'default_body': default_body,
                'default_to': default_to,
                'default_cc': cc,
            })

        # Générer le PDF (identique à observations_pdf)
        html_string = render_to_string('comptabilite/observations_pdf.html', {
            'patient': patient,
            'observations': observations,
            'user': request.user,
        })
        css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')
        base_url = request.build_absolute_uri('/') if settings.DEBUG else f'file://{settings.BASE_DIR}/'
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf(
            stylesheets=[CSS(filename=css_path)]
        )

        email = build_email(
            subject=subject,
            body=body,
            to=to_list or ([request.user.email] if request.user.email else []),
        )
        # Ajout des destinataires en copie si fournis
        if cc_list:
            email.cc = cc_list
        email.attach(f"observations_{dn}.pdf", pdf_bytes, 'application/pdf')
        email.send()
        success_msg = f"PDF des observations envoyé à: {', '.join(to_list)}"
        if cc_list:
            success_msg += f" (Cc: {', '.join(cc_list)})"
        messages.success(request, success_msg)
        return redirect('patients_detail', dn=dn)

    # GET: formulaire par défaut
    default_subject = f"Observations – {patient.nom} {patient.prenom} (DN {dn})"
    default_body = "Bonjour,\n\nVeuillez trouver ci-joint le PDF des observations.\n\nBien cordialement,\nDr. Bronstein"
    # Par défaut: To = le médecin + l'utilisateur si possible; Cc = secrétariat (sans doublons)
    _cands = [request.user.email if getattr(request.user, 'email', '') else None, 'docteur@bronstein.fr']
    _cands = [x for x in _cands if x]
    default_to = ", ".join(list(dict.fromkeys([x.lower() for x in _cands])))
    return render(request, 'comptabilite/observations_email.html', {
        'patient': patient,
        'dn': dn,
        'default_subject': default_subject,
        'default_body': default_body,
        'default_to': default_to,
        'default_cc': 'secretariat@bronstein.fr',
    })


@login_required
def observation_send_email(request, pk: int):
    """Envoie par e‑mail le PDF d'une observation unique (pk)."""
    obs = get_object_or_404(Observation, pk=pk)
    # Identité patient
    fact = Facturation.objects.filter(dn=obs.dn).order_by('-date_acte').first()
    class P: pass
    patient = P()
    if fact:
        patient.dn = fact.dn
        patient.nom = fact.nom
        patient.prenom = fact.prenom
        patient.date_naissance = fact.date_naissance
    else:
        patient.dn = obs.dn
        patient.nom = obs.nom
        patient.prenom = obs.prenom
        patient.date_naissance = obs.date_naissance

    if request.method == 'POST':
        to = (request.POST.get('to') or '').strip()
        cc = (request.POST.get('cc') or '').strip()
        subject = (request.POST.get('subject') or '').strip() or f"Observation du {obs.date_observation.strftime('%d/%m/%Y')} – {patient.nom} {patient.prenom} (DN {patient.dn})"
        body = (request.POST.get('body') or '').strip() or "Bonjour,\n\nVeuillez trouver ci-joint l'observation.\n\nBien cordialement,\nDr. Bronstein"
        to_list = [a.strip() for a in to.replace(';', ',').split(',') if a.strip()]
        cc_list = [a.strip() for a in cc.replace(';', ',').split(',') if a.strip()]

        if not to_list:
            messages.error(request, "Veuillez renseigner au moins un destinataire.")
            default_subject = f"Observation du {obs.date_observation.strftime('%d/%m/%Y')} – {patient.nom} {patient.prenom} (DN {patient.dn})"
            default_body = "Bonjour,\n\nVeuillez trouver ci-joint l'observation.\n\nBien cordialement,\nDr. Bronstein"
            _cands = [request.user.email if getattr(request.user, 'email', '') else None, 'docteur@bronstein.fr']
            _cands = [x for x in _cands if x]
            default_to = ", ".join(list(dict.fromkeys([x.lower() for x in _cands])))
            return render(request, 'comptabilite/observations_email.html', {
                'patient': patient,
                'dn': patient.dn,
                'default_subject': default_subject,
                'default_body': default_body,
                'default_to': default_to,
                'default_cc': cc or 'secretariat@bronstein.fr',
            })

        # Générer PDF pour une seule observation
        html_string = render_to_string('comptabilite/observations_pdf.html', {
            'patient': patient,
            'observations': [obs],
            'user': request.user,
        })
        css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')
        pdf_bytes = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
            stylesheets=[CSS(filename=css_path)]
        )
        email = build_email(
            subject=subject,
            body=body,
            to=to_list or ([request.user.email] if getattr(request.user, 'email', '') else []),
        )
        if cc_list:
            email.cc = cc_list
        email.attach(f"observation_{obs.pk}.pdf", pdf_bytes, 'application/pdf')
        email.send()
        msg = f"Observation du {obs.date_observation.strftime('%d/%m/%Y')} envoyée à: {', '.join(to_list)}"
        if cc_list:
            msg += f" (Cc: {', '.join(cc_list)})"
        messages.success(request, msg)
        return redirect('observations_by_dn', dn=patient.dn)

    # GET: formulaire par défaut
    default_subject = f"Observation du {obs.date_observation.strftime('%d/%m/%Y')} – {patient.nom} {patient.prenom} (DN {patient.dn})"
    default_body = "Bonjour,\n\nVeuillez trouver ci-joint l'observation.\n\nBien cordialement,\nDr. Bronstein"
    _cands = [request.user.email if getattr(request.user, 'email', '') else None, 'docteur@bronstein.fr']
    _cands = [x for x in _cands if x]
    default_to = ", ".join(list(dict.fromkeys([x.lower() for x in _cands])))
    return render(request, 'comptabilite/observations_email.html', {
        'patient': patient,
        'dn': patient.dn,
        'default_subject': default_subject,
        'default_body': default_body,
        'default_to': default_to,
        'default_cc': 'secretariat@bronstein.fr',
    })


# ========== Courriers médicaux (liste par DN, CRUD, PDF, Email) ==========

@login_required
def courriers_by_dn(request, dn: str):
    courriers = Courrier.objects.filter(dn=dn).order_by('-date_courrier', '-created_at')
    fact = Facturation.objects.filter(dn=dn).order_by('-date_acte').first()
    if fact:
        patient = {'dn': fact.dn, 'nom': fact.nom, 'prenom': fact.prenom, 'date_naissance': fact.date_naissance}
    else:
        any_obs = Observation.objects.filter(dn=dn).order_by('-date_observation', '-created_at').first()
        patient = {'dn': dn, 'nom': getattr(any_obs, 'nom', ''), 'prenom': getattr(any_obs, 'prenom', ''), 'date_naissance': getattr(any_obs, 'date_naissance', None)}

    return render(request, 'comptabilite/courriers_list.html', {
        'patient': patient,
        'courriers': courriers,
    })


@login_required
def courrier_create(request, dn: str):
    initial = {'dn': dn}
    # Pré-remplir identité depuis dernière facture si dispo
    fact = Facturation.objects.filter(dn=dn).order_by('-date_acte').first()
    if fact:
        initial.update({'nom': fact.nom, 'prenom': fact.prenom, 'date_naissance': fact.date_naissance})
    else:
        any_obs = Observation.objects.filter(dn=dn).order_by('-date_observation', '-created_at').first()
        if any_obs:
            initial.update({'nom': any_obs.nom, 'prenom': any_obs.prenom, 'date_naissance': any_obs.date_naissance})
    # Type de courrier proposé via GET ?type=
    tp = (request.GET.get('type') or '').upper()
    if tp in dict(Courrier.TYPE_CHOICES):
        initial['type_courrier'] = tp
        # Pré-remplir un modèle de texte selon le type choisi
    if tp == 'FOGD':
            # Date du jour et identité déjà pré-remplies via initial
            from django.utils import timezone as _tz
            today = _tz.localdate().strftime('%d/%m/%Y')
            nom = initial.get('nom', '')
            prenom = initial.get('prenom', '')
            dna = initial.get('date_naissance')
            dna_str = dna.strftime('%d/%m/%Y') if dna else ''
            initial['corps'] = (
                "TECHNIQUE :\n"
                "Lieu : Clinique PAOFAI, salle 4, Opérateur : Dr. J-A Bronstein\n"
                "Appareil Fibroscopie N°2 - n° série : 210 31 12\n"
                "Protocole de désinfection conforme à la circulaire DGH/DH n° 96-236 du 12 avril 1996. "
                "L'opérateur s'est assuré de la validation finale du protocole de désinfection : Oui\n"
                "Préparation : Xylocaïne Gel\n"
                "Tolérance : Bonne\n\n"
                "INDICATION : \n\n"
                "RESULTATS :\n\n"
                "Œsophage : \n"
                "- Muqueuse œsophagienne normale sur toute sa longueur\n"
                "- Ligne Z : 40 cm des arcades dentaires\n\n"
                "Cardia : Cardia anatomique à 40 cm des arcades dentaires\n\n"
                "Fundus : \n\n"
                "Antre : Normal\n\n"
                "Pylore : Normal\n\n"
                "Bulbe : Normal\n\n"
                "Duodénum : Normal\n\n"
                "Biopsies : Oui / Non\n\n"
                "CONCLUSION : Fibroscopie oeso-gastro-duodénale normale\n"
            )
    elif tp == 'COLO':
            from django.utils import timezone as _tz
            today = _tz.localdate().strftime('%d/%m/%Y')
            nom = initial.get('nom', '')
            prenom = initial.get('prenom', '')
            dna = initial.get('date_naissance')
            dna_str = dna.strftime('%d/%m/%Y') if dna else ''
            initial['corps'] = (
                "TECHNIQUE :\n"
                "Lieu : Clinique PAOFAI, salle 4, Opérateur : Dr. J-A Bronstein\n"
                "Appareil Fibroscopie N°2 - n° série : 210 31 12\n"
                "Protocole de désinfection conforme à la circulaire DGH/DH n° 96-236 du 12 avril 1996. "
                "L'opérateur s'est assuré de la validation finale du protocole de désinfection : Oui\n"
                "Préparation : 3 litres de FORTRANS\n"
                "Score de BOSTON : Colon droit :  3  // Colon transverse :  3  // Colon Gauche - Rectum : 3  // Total : 9 \n"
                "Tolérance : Bonne\n\n"
                "INDICATION : \n\n"
                "RESULTATS :\n\n"
                "Progression facile dans un colon bien préparé jusque dans le caecum\n\n"
                "Toucher rectal : normal\n\n"
                "Marge anale : normale\n\n"
                "Canal anal : normal\n\n"
                "Rectum : normal\n\n"
                "Sigmoide : normal\n\n"
                "Colon gauche : normal\n\n"
                "Angle gauche : normal\n\n"
                "Colon transverse : normal\n\n"
                "Angle droit : normal\n\n"
                "Colon droit : normal\n\n"
                "Caecum : normal\n\n"
                "Valvule de Bauhin : normal\n\n"
                "Iléon terminal : normal\n\n"
                "CONCLUSION : Coloscopie totale normale\n"
            )
    elif tp == 'ECHO':
            from django.utils import timezone as _tz
            today = _tz.localdate().strftime('%d/%m/%Y')
            nom = (initial.get('nom') or '').strip()
            prenom = (initial.get('prenom') or '').strip()
            dna = initial.get('date_naissance')
            dna_str = dna.strftime('%d/%m/%Y') if dna else ''
            initial['corps'] = (
                "TECHNIQUE :\n"
                "Lieu : Clinique PAOFAI, Opérateur : Dr. J-A Bronstein\n"
                "Appareil : Sonde convexe haute résolution\n"
                "Tolérance : Bonne\n\n"
                "INDICATION : \n\n"
                "RÉSULTATS :\n\n"
                "Foie : Taille et échostructure normales, sans lésion focale visible\n\n"
                "Voies biliaires : Non dilatées\n\n"
                "Vésicule biliaire : Parois fines, pas de lithiases\n\n"
                "Pancréas : Visualisation correcte, échostructure conservée\n\n"
                "Rate : Taille normale, homogène\n\n"
                "Reins : Taille conservée, pas d'hydronéphrose, pas de lésion focale\n\n"
                "Aorte et VCI : Calibre normal, pas d'anévrysme\n\n"
                "Vessie : Paroi fine, contenu anéchogène\n\n"
                "Prostate/Utérus-Annexes : Aspect dans les limites de la normale (selon contexte)\n\n"
                "CONCLUSION : Échographie abdominale normale\n"
            )
    elif tp == 'SYN':
            from django.utils import timezone as _tz
            today = _tz.localdate().strftime('%d/%m/%Y')
            nom = (initial.get('nom') or '').strip()
            prenom = (initial.get('prenom') or '').strip()
            dna = initial.get('date_naissance')
            dna_str = dna.strftime('%d/%m/%Y') if dna else 'XX/XX/XXXX'
            initial['corps'] = (
                "Cher collègue,\n\n"
                f"M. (Mme) {nom} {prenom}, né(e) le {dna_str}, a été hospitalisé(e) dans le service des hospitalisations ambulatoires le {today} pour dépistage endoscopique sous AG.\n\n"
                "La gastroscopie est normale. Il n'y a pas d'anomalie des muqueuses duodénale, antrale, fundique et œsophagienne. Aucune biopsie n'a été réalisée.\n\n"
                "La coloscopie est complète, menée jusque dans le caecum. Le colon est correctement préparé, et toute la muqueuse est examinée. Au retrait du coloscope, il n'a pas été mis en évidence d'anomalie de la muqueuse (caecum, colon droit, colon transverse, colon gauche, sigmoïde et rectum). Il n'y a pas lieu d'envisager un contrôle avant cinq ans, sauf si apparaissent entre‑temps des rectorragies, des douleurs abdominales, ou un trouble du transit.\n\n"
                "— Variante si polypes —\n"
                "La coloscopie est complète, menée jusque dans le caecum. Le colon est correctement préparé, et toute la muqueuse est examinée. Au retrait du coloscope, on enlève X polype(s) à la pince biopsique. Il n'y a pas lieu d'envisager un contrôle avant cinq ans, sauf si apparaissent entre‑temps des rectorragies, des douleurs abdominales, ou un trouble du transit.\n\n"
                "Je te remercie de ta confiance et je reste à ta disposition.\n"
            )
    elif tp == 'CONS':
            from django.utils import timezone as _tz
            today = _tz.localdate().strftime('%d/%m/%Y')
            nom = (initial.get('nom') or '').strip()
            prenom = (initial.get('prenom') or '').strip()
            dna = initial.get('date_naissance')
            dna_str = dna.strftime('%d/%m/%Y') if dna else 'XX/XX/XXXX'
            initial['corps'] = (
                "\n\n"
                f"Patient : {nom} {prenom}, né(e) le {dna_str} (DN {dn})\n\n"
                "Examen prévu : [sélection ci‑dessous]\n"
                "Date présumée de l'examen : JJ/MM/AAAA\n\n"
                "Information délivrée au patient :\n"
                "- Bénéfices attendus de l'examen\n"
                "- Risques potentiels (hémorragie, perforation, infection, effets de l'anesthésie)\n"
                "- Alternatives possibles\n"
                "- Réponses aux questions posées\n\n"
                "Consentement :\n"
                "Le patient déclare avoir reçu une information claire, loyale et appropriée, et consent à la réalisation de l'examen sus‑mentionné.\n\n"
                f"Date : {today}\\n"
                "Signature du patient : __________________________\\n\\n"
            )
    elif tp == 'ATRE':
            from django.utils import timezone as _tz
            today = _tz.localdate().strftime('%d/%m/%Y')
            nom = (initial.get('nom') or '').strip()
            prenom = (initial.get('prenom') or '').strip()
            dna = initial.get('date_naissance')
            dna_str = dna.strftime('%d/%m/%Y') if dna else 'XX/XX/XXXX'
            initial['corps'] = (
                "\n\n"
                f"Je soussigné Docteur Jean-Ariel BRONSTEIN, certifie que l'état de santé de M.(Mme) {nom} {prenom}, né(e) le {dna_str}, N° DN {dn} l'autorise à retourner vers son domicile par voie aérienne [accompagné(e) / non accompagné(e)].\n\n"
                "Certificat réalisé à la demande de l'intéressé(e) pour servir et remis pour faire valoir ce que de droit.\n\n"
                f"Date : {today}\n\n"
            )
    form = CourrierForm(request.POST or None, request.FILES or None, initial=initial)
    if form.is_valid():
        c = form.save()
        # Mettre à jour les légendes existantes par position (avant suppressions)
        try:
            for p in list(c.photos.all()[:4]):
                new_leg = (request.POST.get(f'legend{p.position}') or '').strip()
                if new_leg != p.legend:
                    p.legend = new_leg
                    p.save(update_fields=['legend'])
        except Exception:
            pass
        # Photos upload (max 4) for eligible types
        if c.type_courrier in ('FOGD', 'COLO', 'ECHO', 'SYN'):
            for i in range(1,5):
                f = request.FILES.get(f'photo{i}')
                if f:
                    from .models import CourrierPhoto
                    legend = (request.POST.get(f'legend{i}') or '').strip()
                    CourrierPhoto.objects.create(courrier=c, image=f, position=i, legend=legend)
        # Création éventuelle d'un rappel anapath si biopsie réalisée (FOGD/COLO)
        try:
            tp_courrier = c.type_courrier
            if tp_courrier in ('FOGD', 'COLO') and request.POST.get('biopsie_realisee'):
                from django.utils import timezone as _tz
                from .models import BiopsyReminder
                dest = (request.POST.get('biopsie_destination') or '').upper()
                if dest in dict(BiopsyReminder.DEST_CHOICES):
                    exam_date = _tz.localdate()
                    send_on = exam_date + _tz.timedelta(days=28)
                    BiopsyReminder.objects.create(
                        dn=c.dn,
                        nom=c.nom,
                        prenom=c.prenom,
                        date_naissance=c.date_naissance,
                        type_examen=tp_courrier,
                        destination=dest,
                        exam_date=exam_date,
                        send_on=send_on,
                    )
        except Exception:
            # On ne bloque pas la création du courrier si le rappel échoue
            pass
        return redirect('courriers_by_dn', dn=c.dn)
    return render(request, 'comptabilite/courrier_form.html', {'form': form, 'mode': 'create', 'dn': dn})


@login_required
def courrier_update(request, pk: int):
    c = get_object_or_404(Courrier, pk=pk)
    form = CourrierForm(request.POST or None, request.FILES or None, instance=c)
    if form.is_valid():
        c = form.save()
        # Mettre à jour les légendes existantes par position
        for p in list(c.photos.all()[:4]):
            new_leg = (request.POST.get(f'legend{p.position}') or '').strip()
            if new_leg != p.legend:
                p.legend = new_leg
                p.save(update_fields=['legend'])
        # Handle deletions
        for photo in list(c.photos.all()):
            if request.POST.get(f'delete_photo_{photo.id}'):
                photo.delete()
        # New uploads (replace vacant positions up to 4)
        if c.type_courrier in ('FOGD', 'COLO', 'ECHO', 'SYN'):
            existing_positions = {p.position for p in c.photos.all()}
            next_pos = 1
            for i in range(1,5):
                f = request.FILES.get(f'photo{i}')
                if f:
                    # assign requested slot i if free else first free
                    pos = i if i not in existing_positions else None
                    if pos is None:
                        # find first free
                        for cand in range(1,5):
                            if cand not in existing_positions:
                                pos = cand
                                break
                    if pos and pos <=4:
                        from .models import CourrierPhoto
                        legend = (request.POST.get(f'legend{pos}') or '').strip()
                        CourrierPhoto.objects.create(courrier=c, image=f, position=pos, legend=legend)
                        existing_positions.add(pos)
        return redirect('courriers_by_dn', dn=c.dn)
    return render(request, 'comptabilite/courrier_form.html', {'form': form, 'mode': 'update', 'dn': c.dn, 'courrier': c})


@login_required
def courrier_delete(request, pk: int):
    c = get_object_or_404(Courrier, pk=pk)
    if request.method == 'POST':
        dn = c.dn
        c.delete()
        return redirect('courriers_by_dn', dn=dn)
    return render(request, 'comptabilite/courrier_confirm_delete.html', {'courrier': c})


@login_required
def courrier_pdf(request, pk: int):
    c = get_object_or_404(Courrier, pk=pk)
    # Construit un objet patient simple
    class P: pass
    patient = P()
    patient.dn = c.dn
    patient.nom = c.nom
    patient.prenom = c.prenom
    patient.date_naissance = c.date_naissance

    template_name = 'comptabilite/courrier_pdf.html'
    if c.type_courrier == 'FOGD':
        template_name = 'comptabilite/courrier_fogd_pdf.html'
    elif c.type_courrier == 'COLO':
        template_name = 'comptabilite/courrier_colo_pdf.html'
    elif c.type_courrier == 'ECHO':
        template_name = 'comptabilite/courrier_echo_pdf.html'
    elif c.type_courrier == 'SYN':
        template_name = 'comptabilite/courrier_syn_pdf.html'
    elif c.type_courrier == 'CONS':
        template_name = 'comptabilite/courrier_cons_pdf.html'
    elif c.type_courrier == 'ATRE':
        template_name = 'comptabilite/courrier_attestation_retour_pdf.html'

    # Photos associées (max 4)
    photos = list(c.photos.all()[:4])
    logo_abs = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
    html_string = render_to_string(template_name, {
        'patient': patient,
        'courrier': c,
        'user': request.user,
        'photos': photos,
        'use_file_src': not settings.DEBUG,
        'embed_images': not settings.DEBUG,
        'logo_abs': logo_abs,
    })
    css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')
    base_url = request.build_absolute_uri('/') if settings.DEBUG else f'file://{settings.BASE_DIR}/'
    pdf = HTML(string=html_string, base_url=base_url).write_pdf(
        stylesheets=[CSS(filename=css_path)]
    )
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="courrier_{c.pk}.pdf"'
    return response


@login_required
def courrier_send_email(request, pk: int):
    c = get_object_or_404(Courrier, pk=pk)
    class P: pass
    patient = P()
    patient.dn = c.dn
    patient.nom = c.nom
    patient.prenom = c.prenom
    patient.date_naissance = c.date_naissance

    type_label = c.type_label()
    date_str = c.date_courrier.strftime('%d/%m/%Y') if c.date_courrier else ''
    if request.method == 'POST':
        to = (request.POST.get('to') or '').strip()
        cc = (request.POST.get('cc') or '').strip()
        subject = (request.POST.get('subject') or '').strip() or f"{type_label} – {patient.nom} {patient.prenom} (DN {patient.dn})"
        body = (request.POST.get('body') or '').strip() or "Bonjour,\n\nVeuillez trouver ci-joint le courrier.\n\nBien cordialement,\nDr. Bronstein"
        to_list = [a.strip() for a in to.replace(';', ',').split(',') if a.strip()]
        cc_list = [a.strip() for a in cc.replace(';', ',').split(',') if a.strip()]

        if not to_list:
            messages.error(request, "Veuillez renseigner au moins un destinataire.")
            default_subject = f"{type_label} – {date_str} – {patient.nom} {patient.prenom} (DN {patient.dn})"
            default_body = "Bonjour,\n\nVeuillez trouver ci-joint le courrier.\n\nBien cordialement,\nDr. Bronstein"
            # Destinataires par défaut selon type
            if c.type_courrier in ('FOGD', 'COLO', 'ECHO', 'CONS'):
                default_to = 'secretariat@bronstein.fr'
                default_cc = cc or 'lwuilmet@polyclinique-paofai.pf'
            elif c.type_courrier == 'SYN':
                default_to = ''  # pas de destinataire principal
                # deux adresses en copie
                default_cc = cc or 'secretariat@bronstein.fr, lwuilmet@polyclinique-paofai.pf'
            else:
                default_to = ''
                default_cc = cc or ''
            return render(request, 'comptabilite/courrier_email.html', {
                'patient': patient,
                'courrier': c,
                'default_subject': default_subject,
                'default_body': default_body,
                'default_to': default_to,
                'default_cc': default_cc,
            })

        # Choix du template PDF identique à l'aperçu PDF
        template_name = 'comptabilite/courrier_pdf.html'
        if c.type_courrier == 'FOGD':
            template_name = 'comptabilite/courrier_fogd_pdf.html'
        elif c.type_courrier == 'COLO':
            template_name = 'comptabilite/courrier_colo_pdf.html'
        elif c.type_courrier == 'ECHO':
            template_name = 'comptabilite/courrier_echo_pdf.html'
        elif c.type_courrier == 'SYN':
            template_name = 'comptabilite/courrier_syn_pdf.html'
        elif c.type_courrier == 'CONS':
            template_name = 'comptabilite/courrier_cons_pdf.html'
        elif c.type_courrier == 'ATRE':
            template_name = 'comptabilite/courrier_attestation_retour_pdf.html'

        photos = list(c.photos.all()[:4])
        logo_abs = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
        html_string = render_to_string(template_name, {
            'patient': patient,
            'courrier': c,
            'user': request.user,
            'photos': photos,
            'use_file_src': not settings.DEBUG,
            'embed_images': not settings.DEBUG,
            'logo_abs': logo_abs,
        })
        css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')
        base_url = request.build_absolute_uri('/') if settings.DEBUG else f'file://{settings.BASE_DIR}/'
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf(
            stylesheets=[CSS(filename=css_path)]
        )
        email = build_email(
            subject=subject,
            body=body,
            to=to_list or ([request.user.email] if getattr(request.user, 'email', '') else []),
        )
        if cc_list:
            email.cc = cc_list
        email.attach(f"courrier_{c.pk}.pdf", pdf_bytes, 'application/pdf')
        email.send()
        msg = f"Courrier envoyé à: {', '.join(to_list)}"
        if cc_list:
            msg += f" (Cc: {', '.join(cc_list)})"
        messages.success(request, msg)
        return redirect('courriers_by_dn', dn=patient.dn)

    default_subject = f"{type_label} – {date_str} – {patient.nom} {patient.prenom} (DN {patient.dn})"
    default_body = "Bonjour,\n\nVeuillez trouver ci-joint le courrier.\n\nBien cordialement,\nDr. Bronstein"
    # Définir les destinataires par défaut selon le type
    if c.type_courrier in ('FOGD', 'COLO', 'ECHO', 'CONS', 'ATRE'):
        default_to = 'secretariat@bronstein.fr'
        default_cc = 'lwuilmet@polyclinique-paofai.pf'
    elif c.type_courrier == 'SYN':
        default_to = ''  # pas de destinataire principal
        default_cc = 'secretariat@bronstein.fr, lwuilmet@polyclinique-paofai.pf'
    else:
        default_to = ''
        default_cc = ''
    return render(request, 'comptabilite/courrier_email.html', {
        'patient': patient,
        'courrier': c,
        'default_subject': default_subject,
        'default_body': default_body,
        'default_to': default_to,
        'default_cc': default_cc,
    })


# ========== Patients (référentiel central) ==========

@login_required
def patients_list(request):
    """Liste des patients avec recherche (DN, nom, prénom, date naissance)."""
    q = (request.GET.get('q') or '').strip()
    qs = Patient.objects.all()
    if q:
        d = parse_flex_date(q)
        filt = (
            Q(dn__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q)
        )
        if d:
            filt |= Q(date_naissance=d)
        qs = qs.filter(filt)
    qs = qs.order_by('nom', 'prenom')
    return render(request, 'comptabilite/patient_list.html', {
        'patients': qs,
        'q': q,
    })


@login_required
def patients_detail(request, dn: str):
    """Fiche patient: infos d'identité et liste des observations, avec actions rapides."""
    # Tente de charger le patient; si absent, essaie de le créer automatiquement
    # depuis la dernière Facturation ou une Observation existante pour ce DN.
    try:
        patient = Patient.objects.get(dn=dn)
    except Patient.DoesNotExist:
        # 1) Source prioritaire: dernière facturation (identité la plus fiable)
        fact = (Facturation.objects
                .filter(dn=dn)
                .order_by('-date_acte', '-id')
                .first())
        if fact:
            patient = Patient.objects.create(
                dn=dn,
                nom=fact.nom or '',
                prenom=fact.prenom or '',
                date_naissance=fact.date_naissance,
            )
            messages.info(request, "Fiche patient créée automatiquement depuis la facturation.")
        else:
            # 2) À défaut, on tente depuis une observation existante
            any_obs = (Observation.objects
                       .filter(dn=dn)
                       .order_by('-date_observation', '-created_at')
                       .first())
            if any_obs:
                patient = Patient.objects.create(
                    dn=dn,
                    nom=any_obs.nom or '',
                    prenom=any_obs.prenom or '',
                    date_naissance=any_obs.date_naissance,
                )
                messages.info(request, "Fiche patient créée automatiquement depuis les observations.")
            else:
                # 3) Rien trouvé → message et redirection douce vers la liste
                messages.warning(request, "Aucune fiche patient ni donnée liée trouvée pour ce DN.")
                return redirect('patients_list')
    observations = Observation.objects.filter(dn=dn).order_by('-date_observation', '-created_at')

    # Dernière facturation utile pour info (facultatif)
    last_fact = (Facturation.objects
                 .filter(dn=dn)
                 .order_by('-date_acte', '-id')
                 .first())
    # Historique de facturation
    facturations = (Facturation.objects
                    .filter(dn=dn)
                    .select_related('code_acte')
                    .order_by('-date_acte', '-id'))

    ctx = {
        'patient': patient,
        'observations': observations,
        'last_fact': last_fact,
        'facturations': facturations,
    }
    return render(request, 'comptabilite/patient_detail.html', ctx)


@login_required
def patient_update(request, dn: str):
    patient = get_object_or_404(Patient, dn=dn)
    form = PatientForm(request.POST or None, instance=patient)
    if form.is_valid():
        form.save()
        return redirect('patients_detail', dn=patient.dn)
    return render(request, 'comptabilite/patient_form.html', {'form': form, 'patient': patient})


@login_required
def patient_delete(request, dn: str):
    patient = get_object_or_404(Patient, dn=dn)
    if request.method == 'POST':
        patient.delete()
        return redirect('patients_list')
    return render(request, 'comptabilite/patient_confirm_delete.html', {'patient': patient})


@login_required
def export_all_data_excel(request):
    """
    Exporte toutes les données de facturation en format Excel
    avec les colonnes : date, nom, total, lieu, mois, annee, code_reel
    """
    if openpyxl is None:
        return HttpResponse("openpyxl non installé – export Excel indisponible.", status=500)
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Données Facturation"
    
    # Définir les en-têtes de colonnes
    headers = ['date', 'nom', 'total', 'lieu', 'mois', 'annee', 'code_reel']
    ws.append(headers)
    
    # Styliser les en-têtes
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer toutes les facturations
    facturations = Facturation.objects.all().order_by('-date_acte')
    
    # Ajouter les données
    for facturation in facturations:
        # Préparer les données de la ligne
        date_acte = facturation.date_acte
        nom_complet = f"{facturation.nom} {facturation.prenom}"
        total = float(facturation.total_acte) if facturation.total_acte else 0
        lieu = facturation.lieu_acte
        mois = date_acte.month if date_acte else ""
        annee = date_acte.year if date_acte else ""
        code_reel = ""
        if facturation.code_acte and facturation.code_acte.code_reel:
            code_reel = facturation.code_acte.code_reel
        
        # Ajouter la ligne au fichier Excel
        ws.append([
            date_acte.strftime('%d/%m/%Y') if date_acte else '',
            nom_complet,
            total,
            lieu,
            mois,
            annee,
            code_reel
        ])
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 12,  # Date
        'B': 25,  # Nom
        'C': 12,  # Total
        'D': 15,  # Lieu
        'E': 8,   # Mois
        'F': 8,   # Année
        'G': 15   # Code reel
    }
    
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    
    # Formater la colonne des totaux
    from openpyxl.styles import NamedStyle
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = '#,##0" XPF"'
    
    for row in range(2, len(facturations) + 2):
        ws[f'C{row}'].style = currency_style
    
    # Créer la réponse HTTP avec le fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nom du fichier avec la date actuelle
    from datetime import datetime
    today = datetime.now().strftime('%Y%m%d')
    filename = f'facturation_complete_{today}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le fichier Excel dans la réponse
    wb.save(response)
    
    return response


@login_required
def export_filtered_data_excel(request):
    """
    Exporte les données de facturation avec filtres optionnels
    Paramètres GET acceptés:
    - date_debut: date de début (format YYYY-MM-DD)
    - date_fin: date de fin (format YYYY-MM-DD) 
    - lieu: filtrer par lieu (Cabinet/Clinique)
    """
    # Récupérer les paramètres de filtre
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    lieu_filtre = request.GET.get('lieu')
    
    if openpyxl is None:
        return HttpResponse("openpyxl non installé – export Excel indisponible.", status=500)
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Données Facturation Filtrées"
    
    # Définir les en-têtes de colonnes
    headers = ['date', 'nom', 'total', 'lieu', 'mois', 'annee', 'code_reel']
    ws.append(headers)
    
    # Styliser les en-têtes
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Construire la requête avec les filtres
    queryset = Facturation.objects.all()
    
    if date_debut:
        from datetime import datetime
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
        queryset = queryset.filter(date_acte__gte=date_debut_obj)
    
    if date_fin:
        from datetime import datetime
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        queryset = queryset.filter(date_acte__lte=date_fin_obj)
    
    if lieu_filtre:
        queryset = queryset.filter(lieu_acte=lieu_filtre)
    
    # Ordonner par date décroissante
    facturations = queryset.order_by('-date_acte')
    
    # Ajouter les données
    for facturation in facturations:
        # Préparer les données de la ligne
        date_acte = facturation.date_acte
        nom_complet = f"{facturation.nom} {facturation.prenom}"
        total = float(facturation.total_acte) if facturation.total_acte else 0
        lieu = facturation.lieu_acte
        mois = date_acte.month if date_acte else ""
        annee = date_acte.year if date_acte else ""
        code_reel = ""
        if facturation.code_acte and facturation.code_acte.code_reel:
            code_reel = facturation.code_acte.code_reel
        
        # Ajouter la ligne au fichier Excel
        ws.append([
            date_acte.strftime('%d/%m/%Y') if date_acte else '',
            nom_complet,
            total,
            lieu,
            mois,
            annee,
            code_reel
        ])
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 12,  # Date
        'B': 25,  # Nom
        'C': 12,  # Total
        'D': 15,  # Lieu
        'E': 8,   # Mois
        'F': 8,   # Année
        'G': 15   # Code reel
    }
    
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    
    # Formater la colonne des totaux
    from openpyxl.styles import NamedStyle
    currency_style = NamedStyle(name="currency_filtered")
    currency_style.number_format = '#,##0" XPF"'
    
    for row in range(2, len(facturations) + 2):
        ws[f'C{row}'].style = currency_style
    
    # Créer la réponse HTTP avec le fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nom du fichier avec la date actuelle et les filtres
    from datetime import datetime
    today = datetime.now().strftime('%Y%m%d')
    filename_parts = ['facturation']
    
    if date_debut and date_fin:
        filename_parts.append(f'{date_debut}_to_{date_fin}')
    elif date_debut:
        filename_parts.append(f'from_{date_debut}')
    elif date_fin:
        filename_parts.append(f'until_{date_fin}')
    
    if lieu_filtre:
        filename_parts.append(lieu_filtre.lower())
    
    filename_parts.append(today)
    filename = '_'.join(filename_parts) + '.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le fichier Excel dans la réponse
    wb.save(response)
    
    return response


@login_required
def export_excel_page(request):
    """
    Page d'export des données en Excel avec options de filtrage
    """
    return render(request, 'comptabilite/export_excel.html')


# ========== Fiches d'information (liens + PDFs locaux) ==========

def fiches_information(request):
    """Affiche la page des fiches d'information patient.

    - Lien externe vers le site ANGH.
    - Liste des PDFs disponibles localement dans static/fiches/ (si présents).
    """
    external_url = "https://angh.net/pratique-2/fiches-information-patient/"
    fiches_dir = os.path.join(settings.BASE_DIR, 'static', 'fiches')
    local_pdfs = []
    try:
        if os.path.isdir(fiches_dir):
            for fn in sorted(os.listdir(fiches_dir)):
                if fn.lower().endswith('.pdf'):
                    local_pdfs.append(fn)
    except Exception:
        pass

    return render(request, 'comptabilite/fiches_information.html', {
        'external_url': external_url,
        'local_pdfs': local_pdfs,
    })
